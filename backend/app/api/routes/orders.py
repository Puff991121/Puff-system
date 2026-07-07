from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from math import ceil
from pathlib import Path
from typing import Annotated, Literal

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from pydantic import ValidationError
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.core.security import get_current_user
from app.models.user import User
from app.schemas.order import (
    DeletedOrder,
    ImportErrorItem,
    ImportResult,
    OrderCreate,
    OrderPage,
    OrderRead,
    OrderSummary,
    OrderUpdate,
    PaymentMethod,
    ResponseEnvelope,
)
from app.services.orders import (
    apply_sort,
    create_order,
    delete_order,
    filtered_orders_query,
    get_order,
    get_order_summary,
    list_orders,
    new_order,
    update_order,
)

router = APIRouter()
Db = Annotated[Session, Depends(get_db)]
CurrentUser = Annotated[User, Depends(get_current_user)]
MAX_FILE_SIZE = 5 * 1024 * 1024
MAX_IMPORT_ROWS = 1000
HEADERS = ["日期", "作业要求", "模板", "格式", "学校", "价格", "支付方式"]
MIME_TYPES = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/octet-stream",
    "application/zip",
}


def response(data, message: str = "success") -> dict:  # type: ignore[no-untyped-def]
    return {"code": 0, "message": message, "data": data}


def not_found() -> HTTPException:
    return HTTPException(status_code=404, detail="订单不存在")


def validate_date_range(start_date: date | None, end_date: date | None) -> None:
    if start_date and end_date and start_date > end_date:
        raise HTTPException(status_code=400, detail="开始日期不能晚于结束日期")


@router.get("", response_model=ResponseEnvelope[OrderPage], summary="查询订单列表")
def get_orders(
    db: Db,
    current_user: CurrentUser,
    page: Annotated[int, Query(ge=1)] = 1,
    page_size: Annotated[int, Query(ge=1, le=100)] = 10,
    keyword: str | None = None,
    payment_method: PaymentMethod | None = None,
    start_date: date | None = None,
    end_date: date | None = None,
    sort_by: Literal["order_date", "price", "created_at"] = "order_date",
    sort_order: Literal["asc", "desc"] = "desc",
) -> dict:
    validate_date_range(start_date, end_date)
    items, total = list_orders(
        db,
        current_user.id,
        page=page,
        page_size=page_size,
        keyword=keyword,
        payment_method=payment_method,
        start_date=start_date,
        end_date=end_date,
        sort_by=sort_by,
        sort_order=sort_order,
    )
    return response(
        OrderPage(
            items=[OrderRead.model_validate(item) for item in items],
            total=total,
            page=page,
            page_size=page_size,
            total_pages=ceil(total / page_size) if total else 0,
        )
    )


@router.get("/summary", response_model=ResponseEnvelope[OrderSummary], summary="查询订单成交统计")
def order_summary(
    db: Db,
    current_user: CurrentUser,
    reference_date: date | None = None,
) -> dict:
    return response(
        OrderSummary(**get_order_summary(db, current_user.id, reference_date=reference_date))
    )


@router.get("/export", summary="导出订单 Excel")
def export_orders(
    db: Db,
    current_user: CurrentUser,
    keyword: str | None = None,
    payment_method: PaymentMethod | None = None,
    start_date: date | None = None,
    end_date: date | None = None,
    sort_by: Literal["order_date", "price", "created_at"] = "order_date",
    sort_order: Literal["asc", "desc"] = "desc",
) -> StreamingResponse:
    validate_date_range(start_date, end_date)
    statement = filtered_orders_query(
        current_user.id,
        keyword=keyword,
        payment_method=payment_method,
        start_date=start_date,
        end_date=end_date,
    )
    orders = db.scalars(apply_sort(statement, sort_by, sort_order)).all()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "订单"
    sheet.append(HEADERS)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    for order in orders:
        sheet.append(
            [
                order.order_date,
                order.requirement,
                order.template,
                order.format,
                order.school,
                float(order.price),
                order.payment_method,
            ]
        )
        sheet.cell(sheet.max_row, 1).number_format = "yyyy-mm-dd"
        sheet.cell(sheet.max_row, 6).number_format = "0.00"

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f"orders-{date.today():%Y%m%d}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def excel_date(value: object) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return date.fromisoformat(str(value).strip())


def import_error(row: int, field: str, value: object, message: str) -> ImportErrorItem:
    return ImportErrorItem(
        row=row,
        field=field,
        value=str(value) if value is not None else None,
        message=message,
    )


@router.post("/import", response_model=ResponseEnvelope[ImportResult], summary="导入订单 Excel")
async def import_orders(
    db: Db,
    current_user: CurrentUser,
    file: Annotated[UploadFile, File(...)],
) -> dict:
    if Path(file.filename or "").suffix.lower() != ".xlsx":
        raise HTTPException(
            status_code=status.HTTP_415_UNSUPPORTED_MEDIA_TYPE,
            detail="仅支持 .xlsx 文件",
        )
    if file.content_type not in MIME_TYPES:
        raise HTTPException(
            status_code=status.HTTP_415_UNSUPPORTED_MEDIA_TYPE,
            detail="文件 MIME 类型无效",
        )
    content = await file.read(MAX_FILE_SIZE + 1)
    if len(content) > MAX_FILE_SIZE:
        raise HTTPException(
            status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
            detail="文件不能超过 5MB",
        )

    try:
        workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail="无法读取 Excel 文件") from exc
    sheet = workbook.active
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    if headers != HEADERS:
        raise HTTPException(status_code=400, detail=f"Excel 表头必须为：{'、'.join(HEADERS)}")

    rows = [
        row
        for row in sheet.iter_rows(min_row=2, values_only=True)
        if any(value not in (None, "") for value in row)
    ]
    if len(rows) > MAX_IMPORT_ROWS:
        raise HTTPException(status_code=400, detail="单个文件最多允许 1000 条订单")

    imported = []
    errors: list[ImportErrorItem] = []
    for row_number, values in enumerate(rows, start=2):
        try:
            payload = OrderCreate(
                order_date=excel_date(values[0]),
                requirement=values[1],
                template=values[2],
                format=values[3],
                school=values[4],
                price=Decimal(str(values[5])),
                payment_method=values[6],
            )
            order = new_order(current_user.id, payload)
            db.add(order)
            imported.append(order)
        except (ValidationError, ValueError, InvalidOperation) as exc:
            field = "数据"
            if isinstance(exc, ValidationError) and exc.errors():
                field_name = str(exc.errors()[0]["loc"][0])
                field_map = dict(zip(OrderCreate.model_fields, HEADERS))
                field = field_map.get(field_name, field)
            errors.append(import_error(row_number, field, values, str(exc)))

    try:
        db.commit()
        for order in imported:
            db.refresh(order)
    except Exception:
        db.rollback()
        raise

    result = ImportResult(
        total_rows=len(rows),
        success_count=len(imported),
        failed_count=len(errors),
        imported_orders=[OrderRead.model_validate(order) for order in imported],
        errors=errors,
    )
    return response(result, "订单导入完成")


@router.post("", status_code=201, response_model=ResponseEnvelope[OrderRead], summary="新增订单")
def add_order(payload: OrderCreate, db: Db, current_user: CurrentUser) -> dict:
    order = create_order(db, current_user.id, payload)
    return response(OrderRead.model_validate(order), "订单新增成功")


@router.get("/{order_id}", response_model=ResponseEnvelope[OrderRead], summary="获取订单详情")
def order_detail(order_id: int, db: Db, current_user: CurrentUser) -> dict:
    order = get_order(db, current_user.id, order_id)
    if order is None:
        raise not_found()
    return response(OrderRead.model_validate(order))


@router.patch("/{order_id}", response_model=ResponseEnvelope[OrderRead], summary="修改订单")
def edit_order(order_id: int, payload: OrderUpdate, db: Db, current_user: CurrentUser) -> dict:
    order = get_order(db, current_user.id, order_id)
    if order is None:
        raise not_found()
    order = update_order(db, order, payload)
    return response(OrderRead.model_validate(order), "订单修改成功")


@router.delete("/{order_id}", response_model=ResponseEnvelope[DeletedOrder], summary="删除订单")
def remove_order(order_id: int, db: Db, current_user: CurrentUser) -> dict:
    order = get_order(db, current_user.id, order_id)
    if order is None:
        raise not_found()
    delete_order(db, order)
    return response(DeletedOrder(deleted_id=order_id), "订单删除成功")
