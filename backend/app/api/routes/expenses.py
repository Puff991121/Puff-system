from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from math import ceil
from pathlib import Path
from typing import Annotated, Literal

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from pydantic import ValidationError
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.core.security import get_current_user
from app.models.user import User
from app.schemas.expense import (
    DeletedExpense,
    ExpenseCreate,
    ExpenseImportError,
    ExpenseImportResult,
    ExpensePage,
    ExpenseRead,
    ExpenseSummary,
    ExpenseUpdate,
)
from app.schemas.order import ResponseEnvelope
from app.services.expenses import (
    create_expense,
    delete_expense,
    filtered_expenses_query,
    get_expense,
    get_summary,
    list_expenses,
    new_expense,
    update_expense,
)

router = APIRouter()
Db = Annotated[Session, Depends(get_db)]
CurrentUser = Annotated[User, Depends(get_current_user)]
HEADERS = ["交易时间", "交易类型", "交易对方", "商品说明", "金额", "支付方式"]
MAX_FILE_SIZE = 5 * 1024 * 1024
MAX_IMPORT_ROWS = 1000


def response(data, message="success") -> dict:
    return {"code": 0, "message": message, "data": data}


def validate_range(start_time, end_time) -> None:
    if start_time and end_time and start_time > end_time:
        raise HTTPException(400, "开始时间不能晚于结束时间")


@router.get("", response_model=ResponseEnvelope[ExpensePage], summary="查询消费记录")
def index(
    db: Db,
    current_user: CurrentUser,
    page: Annotated[int, Query(ge=1)] = 1,
    page_size: Annotated[int, Query(ge=1, le=100)] = 10,
    transaction_type: str | None = None,
    description: str | None = None,
    payment_method: str | None = None,
    start_time: datetime | None = None,
    end_time: datetime | None = None,
    sort_by: Literal["transaction_time", "amount", "created_at"] = "transaction_time",
    sort_order: Literal["asc", "desc"] = "desc",
) -> dict:
    validate_range(start_time, end_time)
    items, total = list_expenses(
        db,
        current_user.id,
        page=page,
        page_size=page_size,
        transaction_type=transaction_type,
        description=description,
        payment_method=payment_method,
        start_time=start_time,
        end_time=end_time,
        sort_by=sort_by,
        sort_order=sort_order,
    )
    return response(
        ExpensePage(
            items=[ExpenseRead.model_validate(x) for x in items],
            total=total,
            page=page,
            page_size=page_size,
            total_pages=ceil(total / page_size) if total else 0,
        )
    )


@router.get("/summary", response_model=ResponseEnvelope[ExpenseSummary], summary="消费统计")
def summary(db: Db, current_user: CurrentUser, reference_date: date | None = None) -> dict:
    return response(
        ExpenseSummary(**get_summary(db, current_user.id, reference_date=reference_date))
    )


@router.get("/export", summary="导出消费记录 Excel")
def export(
    db: Db,
    current_user: CurrentUser,
    transaction_type: str | None = None,
    description: str | None = None,
    payment_method: str | None = None,
    start_time: datetime | None = None,
    end_time: datetime | None = None,
) -> StreamingResponse:
    validate_range(start_time, end_time)
    items = db.scalars(
        filtered_expenses_query(
            current_user.id,
            transaction_type=transaction_type,
            description=description,
            payment_method=payment_method,
            start_time=start_time,
            end_time=end_time,
        ).order_by()
    ).all()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "消费记录"
    sheet.append(HEADERS)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    for item in items:
        sheet.append(
            [
                item.transaction_time,
                item.transaction_type,
                item.counterparty,
                item.description,
                float(item.amount),
                item.payment_method,
            ]
        )
        sheet.cell(sheet.max_row, 1).number_format = "yyyy-mm-dd hh:mm:ss"
        sheet.cell(sheet.max_row, 5).number_format = "0.00"
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="expenses.xlsx"'},
    )


@router.post(
    "/import", response_model=ResponseEnvelope[ExpenseImportResult], summary="导入消费记录 Excel"
)
async def import_file(
    db: Db, current_user: CurrentUser, file: Annotated[UploadFile, File(...)]
) -> dict:
    if Path(file.filename or "").suffix.lower() != ".xlsx":
        raise HTTPException(415, "仅支持 .xlsx 文件")
    content = await file.read(MAX_FILE_SIZE + 1)
    if len(content) > MAX_FILE_SIZE:
        raise HTTPException(413, "文件不能超过 5MB")
    try:
        workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
    except Exception as exc:
        raise HTTPException(400, "无法读取 Excel 文件") from exc
    sheet = workbook.active
    headers = [c.value for c in next(sheet.iter_rows(min_row=1, max_row=1))]
    if headers != HEADERS:
        raise HTTPException(400, f"Excel 表头必须为：{'、'.join(HEADERS)}")
    rows = [
        r
        for r in sheet.iter_rows(min_row=2, values_only=True)
        if any(v not in (None, "") for v in r)
    ]
    if len(rows) > MAX_IMPORT_ROWS:
        raise HTTPException(400, "单个文件最多允许 1000 条记录")
    imported, errors = [], []
    for row_number, values in enumerate(rows, 2):
        try:
            time_value = (
                values[0]
                if isinstance(values[0], datetime)
                else datetime.fromisoformat(str(values[0]))
            )
            payload = ExpenseCreate(
                transaction_time=time_value,
                transaction_type=values[1],
                counterparty=values[2],
                description=values[3],
                amount=Decimal(str(values[4])),
                payment_method=values[5],
            )
            item = new_expense(current_user.id, payload)
            db.add(item)
            imported.append(item)
        except (ValidationError, ValueError, InvalidOperation) as exc:
            errors.append(
                ExpenseImportError(
                    row=row_number, field="数据", value=str(values), message=str(exc)
                )
            )
    db.commit()
    for item in imported:
        db.refresh(item)
    result = ExpenseImportResult(
        total_rows=len(rows),
        success_count=len(imported),
        failed_count=len(errors),
        imported_expenses=[ExpenseRead.model_validate(x) for x in imported],
        errors=errors,
    )
    return response(result, "消费记录导入完成")


@router.post(
    "", status_code=201, response_model=ResponseEnvelope[ExpenseRead], summary="新增消费记录"
)
def create(payload: ExpenseCreate, db: Db, current_user: CurrentUser) -> dict:
    return response(
        ExpenseRead.model_validate(create_expense(db, current_user.id, payload)), "新增成功"
    )


@router.get("/{expense_id}", response_model=ResponseEnvelope[ExpenseRead], summary="消费记录详情")
def detail(expense_id: int, db: Db, current_user: CurrentUser) -> dict:
    item = get_expense(db, current_user.id, expense_id)
    if item is None:
        raise HTTPException(404, "消费记录不存在")
    return response(ExpenseRead.model_validate(item))


@router.patch("/{expense_id}", response_model=ResponseEnvelope[ExpenseRead], summary="修改消费记录")
def edit(expense_id: int, payload: ExpenseUpdate, db: Db, current_user: CurrentUser) -> dict:
    item = get_expense(db, current_user.id, expense_id)
    if item is None:
        raise HTTPException(404, "消费记录不存在")
    return response(ExpenseRead.model_validate(update_expense(db, item, payload)), "修改成功")


@router.delete(
    "/{expense_id}", response_model=ResponseEnvelope[DeletedExpense], summary="删除消费记录"
)
def remove(expense_id: int, db: Db, current_user: CurrentUser) -> dict:
    item = get_expense(db, current_user.id, expense_id)
    if item is None:
        raise HTTPException(404, "消费记录不存在")
    delete_expense(db, item)
    return response(DeletedExpense(deleted_id=expense_id), "删除成功")
