from collections.abc import Generator
from datetime import timedelta
from io import BytesIO

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import Session, sessionmaker
from sqlalchemy.pool import StaticPool

from app.core.database import Base, get_db
from app.core.security import create_access_token
from app.main import app
from app.models import User
from app.services.expenses import now

engine = create_engine("sqlite://", connect_args={"check_same_thread": False}, poolclass=StaticPool)
TestingSession = sessionmaker(bind=engine, expire_on_commit=False)


def override_get_db() -> Generator[Session, None, None]:
    with TestingSession() as db:
        yield db


@pytest.fixture(autouse=True)
def database() -> Generator[None, None, None]:
    app.dependency_overrides[get_db] = override_get_db
    Base.metadata.create_all(engine)
    with TestingSession() as db:
        db.add(User(username="admin", password_hash="unused"))
        db.commit()
    yield
    app.dependency_overrides.clear()
    Base.metadata.drop_all(engine)


@pytest.fixture
def client() -> TestClient:
    return TestClient(app)


@pytest.fixture
def headers() -> dict[str, str]:
    return {"Authorization": f"Bearer {create_access_token('admin')}"}


def payload(**changes):
    data = {
        "transaction_time": "2026-07-08T14:30:00+08:00",
        "transaction_type": "商品采购",
        "counterparty": "晨光文具",
        "description": "打印纸 A4",
        "amount": "39.90",
        "payment_method": "微信",
    }
    data.update(changes)
    return data


def test_expense_crud_filters_and_summary(client: TestClient, headers: dict[str, str]) -> None:
    first = client.post("/api/expenses", headers=headers, json=payload())
    assert first.status_code == 201
    expense = first.json()["data"]
    client.post(
        "/api/expenses",
        headers=headers,
        json=payload(
            transaction_type="餐饮",
            description="团队午餐",
            amount="128.00",
            payment_method="支付宝",
        ),
    )
    listed = client.get(
        "/api/expenses?transaction_type=商品采购&description=打印&payment_method=微信",
        headers=headers,
    )
    assert listed.status_code == 200
    assert listed.json()["data"]["total"] == 1
    changed = client.patch(
        f"/api/expenses/{expense['id']}", headers=headers, json={"amount": "49.90"}
    )
    assert changed.json()["data"]["amount"] == "49.90"
    summary = client.get("/api/expenses/summary", headers=headers).json()["data"]
    assert summary["total_count"] == 2
    assert summary["total_amount"] == "177.90"
    assert "today_amount" in summary
    assert "month_change_rate" in summary
    assert "year_amount" in summary
    assert "year_change_rate" in summary
    assert client.delete(f"/api/expenses/{expense['id']}", headers=headers).status_code == 200


def test_expense_export_and_import(client: TestClient, headers: dict[str, str]) -> None:
    client.post("/api/expenses", headers=headers, json=payload())
    exported = client.get("/api/expenses/export", headers=headers)
    assert exported.status_code == 200
    assert load_workbook(BytesIO(exported.content)).active.max_row == 2
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["交易时间", "交易类型", "交易对方", "商品说明", "金额", "支付方式"])
    sheet.append(["2026-07-08 18:00:00", "服务支出", "云服务商", "服务器续费", 199, "银行卡"])
    content = BytesIO()
    workbook.save(content)
    imported = client.post(
        "/api/expenses/import",
        headers=headers,
        files={
            "file": (
                "expenses.xlsx",
                content.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert imported.status_code == 200
    assert imported.json()["data"]["success_count"] == 1


def test_expense_validation_and_auth(client: TestClient, headers: dict[str, str]) -> None:
    assert client.get("/api/expenses").status_code == 401
    assert (
        client.post("/api/expenses", headers=headers, json=payload(amount="0")).status_code == 422
    )
    assert (
        client.get(
            "/api/expenses?start_time=2026-08-01T00:00:00%2B08:00&end_time=2026-07-01T00:00:00%2B08:00",
            headers=headers,
        ).status_code
        == 400
    )

    custom_values = client.post(
        "/api/expenses",
        headers=headers,
        json=payload(
            transaction_type="充值缴费",
            payment_method="余额宝&支付宝随机抽立减",
        ),
    )
    assert custom_values.status_code == 201
    assert custom_values.json()["data"]["transaction_type"] == "充值缴费"
    assert custom_values.json()["data"]["payment_method"] == "余额宝&支付宝随机抽立减"


def test_expense_summary_month_over_month_and_year_over_year(
    client: TestClient, headers: dict[str, str]
) -> None:
    current = now()
    previous_month = current.replace(day=1) - timedelta(days=1)
    previous_year = current.replace(year=current.year - 1)
    for transaction_time, amount in (
        (current, "200.00"),
        (previous_month, "100.00"),
        (previous_year, "50.00"),
    ):
        client.post(
            "/api/expenses",
            headers=headers,
            json=payload(transaction_time=transaction_time.isoformat(), amount=amount),
        )

    summary = client.get("/api/expenses/summary", headers=headers).json()["data"]
    assert summary["month_change_rate"] == "100.00"
    assert summary["year_change_rate"] == "500.00"


def test_expense_summary_uses_queried_month(client: TestClient, headers: dict[str, str]) -> None:
    for transaction_time, amount in (
        ("2024-05-12T10:00:00+08:00", "100.00"),
        ("2024-06-08T10:00:00+08:00", "150.00"),
        ("2024-06-20T10:00:00+08:00", "50.00"),
        ("2024-07-01T10:00:00+08:00", "999.00"),
    ):
        client.post(
            "/api/expenses",
            headers=headers,
            json=payload(transaction_time=transaction_time, amount=amount),
        )

    summary = client.get("/api/expenses/summary?reference_date=2024-06-15", headers=headers).json()[
        "data"
    ]
    assert summary["month_amount"] == "200.00"
    assert summary["month_count"] == 2
    assert summary["month_change_rate"] == "100.00"
